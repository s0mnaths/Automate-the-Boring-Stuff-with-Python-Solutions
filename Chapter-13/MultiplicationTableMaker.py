import openpyxl, sys
from openpyxl.utils import get_column_letter

wb=openpyxl.Workbook()
sheet=wb.active
bold = openpyxl.styles.Font(bold=True)
N=int(sys.argv[1])
sheet.cell(row=1,column=1).value=''
for r in range(1,N+2):
    for c in range(1,N+2):
        if r==1 and c==1:
            sheet.cell(row=r,column=c).value=''
        elif r==1:
            sheet.cell(row=r,column=c).value=c-1
            sheet.cell(row=r,column=c).font=bold
        elif c==1:
            sheet.cell(row=r,column=c).value=r-1
            sheet.cell(row=r,column=c).font=bold
        else:
            sheet.cell(row=r,column=c).value=(r-1)*(c-1)
wb.save('MultiplicationTable.xlsx')