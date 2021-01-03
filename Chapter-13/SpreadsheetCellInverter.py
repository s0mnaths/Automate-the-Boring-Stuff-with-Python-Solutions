import openpyxl,sys

file = sys.argv[1]

wb = openpyxl.load_workbook(file)
wbNew = openpyxl.Workbook()

sheetOld = wb.active
sheetNew = wbNew['Sheet']

row=max(sheetOld.max_row,sheetOld.max_column)
col=max(sheetOld.max_row,sheetOld.max_column)
for r in range(1,row+1):
    for c in range(1,col+1):
        sheetNew.cell(row=r,column=c).value = sheetOld.cell(row=c,column=r).value
       

wbNew.save('example2.xlsx')