#USAGE: Run in command line--> python3 ch-13e.py <excel file name>
                                                #excel file should be in same directory as this script
import openpyxl,sys

wb=openpyxl.load_workbook(sys.argv[1])
sheet=wb.active
row=sheet.max_row
col=sheet.max_column
for c in range(1,col+1):
    fileName='excelToTextCol'+str(c)+'.txt'
    file=open(fileName,'w')
    for r in range(1,row+1):
        file.write(sheet.cell(row=r,column=c).value)
    file.close()