import openpyxl,sys

file = sys.argv[3]

wb = openpyxl.load_workbook(file)
wbNew = openpyxl.Workbook()

N = int(sys.argv[1])
M = int(sys.argv[2])

sheetOld = wb.active
sheetNew = wbNew['Sheet']

for rowObj in sheetOld.rows:
    for cellObj in rowObj:
       
        if cellObj.row<N:
            sheetNew.cell(row=cellObj.row,column=int(cellObj.column)).value = cellObj.value
        else:
            sheetNew.cell(row=cellObj.row+M,column=int(cellObj.column)).value = cellObj.value

wbNew.save(file)